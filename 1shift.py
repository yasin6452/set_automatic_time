from openpyxl import load_workbook

# بارگذاری فایل اکسل
file_path = "ttttttttttttyuu.xlsx"  # مسیر فایل خود را وارد کنید
wb = load_workbook(file_path)
ws = wb.active

# ذخیره سلول‌های ترکیب‌شده
merged_ranges = list(ws.merged_cells.ranges)

# شکستن سلول‌های ترکیب‌شده
for merged_range in merged_ranges:
    ws.unmerge_cells(str(merged_range))  # حذف ترکیب سلول‌ها

# دریافت نام ستون‌ها از سطر اول
header = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]

# جابجایی ستون‌های ورود و خروج (مثال: ورود در ستون AA و خروج در ستون AB)
for row in range(2, ws.max_row + 1):  # از ردیف 2 تا انتها
    # ذخیره مقادیر فعلی ستون‌ها
    entry_value = ws.cell(row=row, column=27).value  # AA = ورود
    exit_value = ws.cell(row=row, column=28).value   # AB = خروج
    
    # جابجایی مقادیر
    ws.cell(row=row, column=27).value = exit_value   # ستون AA (ورود) می‌شود خروج
    ws.cell(row=row, column=28).value = entry_value  # ستون AB (خروج) می‌شود ورود

# جابجایی سایر ستون‌ها به عقب (شیفت ستون‌ها)
for row in range(2, ws.max_row + 1):  # از ردیف 2 تا انتها
    for col in range(29, ws.max_column + 1):  # از ستون 29 به بعد (یعنی ستون‌های بعد از AB)
        # انتقال داده‌ها به عقب
        current_value = ws.cell(row=row, column=col).value
        ws.cell(row=row, column=col + 2).value = current_value  # دو ستون به عقب

# ذخیره فایل با نام جدید
output_path = file_path.replace(".xlsx", "_modified.xlsx")
wb.save(output_path)

print("ستون‌ها جابجا شدند و فایل ذخیره شد در:", output_path)
