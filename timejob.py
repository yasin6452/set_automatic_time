from openpyxl import load_workbook
from datetime import time, timedelta
import random
import os
import shutil

print("🔍 برنامه پرکننده لیست کارکرد - نسخه با دقیقه‌های 30، 40، 50 (7:30 تا 8:00)")
print("=" * 60)

# دریافت فایل
file_path = input("مسیر کامل فایل اکسل: ").strip().strip('"')

if not os.path.exists(file_path):
    print(f"❌ فایل پیدا نشد: {file_path}")
    exit()

print(f"✅ فایل پیدا شد: {os.path.basename(file_path)}")

# ایجاد پشتیبان
backup_path = file_path.replace('.xlsx', '_backup.xlsx')
shutil.copy2(file_path, backup_path)
print(f"📋 پشتیبان گرفته شد: {os.path.basename(backup_path)}")

try:
    # بارگذاری فایل
    wb = load_workbook(file_path)
    
    # انتخاب شیت صحیح
    if 'مریم ایرانمنش' in wb.sheetnames:
        ws = wb['مریم ایرانمنش']
        print(f"✅ شیت 'مریم ایرانمنش' انتخاب شد")
    else:
        ws = wb.active
        print(f"⚠️ شیت فعال: {ws.title}")
    
    print(f"📊 ابعاد شیت: {ws.max_row} سطر × {ws.max_column} ستون")
    
    # بررسی ساختار فایل
    print("\n🔍 بررسی ساختار فایل...")
    
    # پیدا کردن ستون‌های کلیدی
    entry_col = None  # ستون ورود
    exit_col = None   # ستون خروج
    work_col = None   # ستون کارکرد
    day_col = None    # ستون روز
    
    # بررسی سطر 7 و 8 برای هدرها
    for col in range(1, ws.max_column + 1):
        val7 = str(ws.cell(row=7, column=col).value or "")
        val8 = str(ws.cell(row=8, column=col).value or "")
        
        if "ترددها" in val8 and "ورود" in val7:
            entry_col = col
            print(f"✅ ستون ورود پیدا شد: {chr(64 + col)} (ستون {col})")
        elif "ترددها" in val8 and "خروج" in val7:
            exit_col = col
            print(f"✅ ستون خروج پیدا شد: {chr(64 + col)} (ستون {col})")
        elif "طول" in val7 or "كاركرد" in val8:
            work_col = col
            print(f"✅ ستون کارکرد پیدا شد: {chr(64 + col)} (ستون {col})")
        elif "روز" in val7 or "روز" in val8:
            day_col = col
            print(f"✅ ستون روز پیدا شد: {chr(64 + col)} (ستون {col})")
    
    # اگر ستون روز پیدا نشد، از ستون 35 استفاده کن
    if not day_col:
        day_col = 35
        print(f"⚠️ ستون روز پیدا نشد، از ستون {day_col} استفاده می‌کنم")
    
    # اگر ستون ورود/خروج پیدا نشدند
    if not entry_col:
        entry_col = 27  # AA
        print(f"⚠️ ستون ورود پیدا نشد، از ستون {entry_col} استفاده می‌کنم")
    
    if not exit_col:
        exit_col = 28  # AB
        print(f"⚠️ ستون خروج پیدا نشد، از ستون {exit_col} استفاده می‌کنم")
    
    if not work_col:
        work_col = 26  # Z
        print(f"⚠️ ستون کارکرد پیدا نشد، از ستون {work_col} استفاده می‌کنم")
    
    print(f"\n🎯 ستون‌های نهایی:")
    print(f"  • ستون روزها: {chr(64 + day_col)} (شماره {day_col})")
    print(f"  • ستون ورود: {chr(64 + entry_col)} (شماره {entry_col})")
    print(f"  • ستون خروج: {chr(64 + exit_col)} (شماره {exit_col})")
    print(f"  • ستون کارکرد: {chr(64 + work_col)} (شماره {work_col})")
    
    # لیست دقیقه‌های مجاز (بین 30 تا 50)
    allowed_minutes = [30, 40, 50]
    print(f"\n⏰ محدوده زمانی ورود: 7:30 تا 8:00")
    print(f"⏰ دقیقه‌های مجاز برای ورود: {allowed_minutes}")
    
    # تایید از کاربر
    confirm = input("\n↵ برای ادامه Enter بزنید (یا 'n' برای لغو): ").strip().lower()
    if confirm == 'n':
        print("❌ عملیات لغو شد.")
        exit()
    
    print("\n🔄 در حال پر کردن داده‌ها...")
    
    filled_days = 0
    fridays = 0
    minute_counts = {30: 0, 40: 0, 50: 0}  # شمارش دقیقه‌ها
    
    # دنباله‌ای از دقیقه‌ها برای ایجاد تنوع
    minute_sequence = []
    days_to_fill = 30 - 4  # تقریباً 26 روز کاری (30 روز کل منهای 4 جمعه)
    
    # توزیع متعادل دقیقه‌ها
    base_count = days_to_fill // len(allowed_minutes)
    remainder = days_to_fill % len(allowed_minutes)
    
    for minute in allowed_minutes:
        count = base_count + (1 if allowed_minutes.index(minute) < remainder else 0)
        minute_sequence.extend([minute] * count)
    
    # کمی به ترتیب شانسی می‌دهیم
    random.shuffle(minute_sequence)
    
    minute_index = 0
    
    for row in range(9, 40):  # سطر 9 تا 39
        day_value = ws.cell(row=row, column=day_col).value
        
        if not day_value:
            continue
            
        day_str = str(day_value).strip()
        
        if day_str == "جمعه":
            # جمعه: پاک کردن
            ws.cell(row=row, column=entry_col).value = None
            ws.cell(row=row, column=exit_col).value = None
            ws.cell(row=row, column=work_col).value = None
            fridays += 1
            print(f"🚫 سطر {row:2d} (جمعه) - پاک شد")
        else:
            # روز کاری: پر کردن
            
            # زمان ورود: 7:XX (XX فقط یکی از مقادیر 30، 40، 50)
            entry_hour = 7
            
            # انتخاب دقیقه از دنباله
            if minute_index < len(minute_sequence):
                entry_minute = minute_sequence[minute_index]
                minute_index += 1
            else:
                # اگر دنباله تمام شد، از بین دقیقه‌ها انتخاب تصادفی
                entry_minute = random.choice(allowed_minutes)
            
            minute_counts[entry_minute] += 1
            
            entry_time = time(entry_hour, entry_minute)
            
            # مدت کار: 8 ساعت و دقیقه اضافه کاری
            work_hours = 8
            
            # دقیقه اضافه کاری (0-30 دقیقه)
            extra_minutes_options = [0, 5, 10, 15, 20, 25, 30]
            work_minutes = random.choice(extra_minutes_options)
            
            # محاسبه زمان خروج
            total_minutes = (entry_hour * 60 + entry_minute) + (work_hours * 60 + work_minutes)
            exit_hour = total_minutes // 60
            exit_minute = total_minutes % 60
            exit_time = time(exit_hour, exit_minute)
            
            # محاسبه کارکرد به ساعت (با یک رقم اعشار)
            total_work_hours = work_hours + (work_minutes / 60)
            
            # پر کردن سلول‌ها
            ws.cell(row=row, column=entry_col).value = entry_time  # زمان ورود
            ws.cell(row=row, column=exit_col).value = exit_time    # زمان خروج
            ws.cell(row=row, column=work_col).value = round(total_work_hours, 1)  # ساعت کارکرد
            
            # فرمت‌دهی
            ws.cell(row=row, column=entry_col).number_format = "HH:MM"
            ws.cell(row=row, column=exit_col).number_format = "HH:MM"
            ws.cell(row=row, column=work_col).number_format = "0.0"
            
            filled_days += 1
            print(f"✅ سطر {row:2d} ({day_str}): {entry_time.strftime('%H:%M')} - {exit_time.strftime('%H:%M')} ({work_hours}:{work_minutes:02d} ساعت)")
    
    # نمایش آمار دقیقه‌ها
    print(f"\n📊 آمار دقیقه‌های ورود:")
    total_uses = sum(minute_counts.values())
    for minute, count in sorted(minute_counts.items()):
        if count > 0:
            percentage = (count / total_uses) * 100
            print(f"  • {minute:2d} دقیقه (7:{minute:02d}): {count} بار ({percentage:.1f}%)")
    
    print(f"\n📊 نتیجه نهایی:")
    print(f"  • {filled_days} روز کاری پر شد")
    print(f"  • {fridays} جمعه پاک شد")
    print(f"  • محدوده ورود: 7:30 تا 8:00")
    print(f"  • مدت کار پایه: 8 ساعت")
    print(f"  • اضافه کاری: 0 تا 30 دقیقه")
    
    # ذخیره فایل
    default_name = os.path.basename(file_path).replace('.xlsx', '_پر شده.xlsx')
    output_name = input(f"\n📝 نام فایل خروجی (Enter برای '{default_name}'): ").strip()
    
    if not output_name:
        output_name = default_name
    elif not output_name.endswith('.xlsx'):
        output_name += '.xlsx'
    
    output_path = os.path.join(os.path.dirname(file_path), output_name)
    
    try:
        wb.save(output_path)
        print(f"\n🎉 فایل با موفقیت ذخیره شد!")
        print(f"📁 مسیر: {output_path}")
        print(f"📏 حجم: {os.path.getsize(output_path):,} بایت")
        
        # نمایش خلاصه
        print(f"\n📋 خلاصه تغییرات:")
        print(f"  1. ساعت ورود: 7:30 تا 8:00")
        print(f"  2. دقیقه ورود: فقط 30، 40 یا 50 دقیقه")
        print(f"  3. مدت کار پایه: 8 ساعت")
        print(f"  4. اضافه کاری: 0، 5، 10، 15، 20، 25 یا 30 دقیقه")
        print(f"  5. روزهای جمعه خالی شدند")
        print(f"  6. فرمت زمان: HH:MM")
        print(f"  7. فرمت کارکرد: 0.0 ساعت")
        
    except Exception as e:
        print(f"❌ خطا در ذخیره فایل: {e}")
        print("⚠️ در حال ذخیره با نام پیش‌فرض...")
        try:
            fallback_path = file_path.replace('.xlsx', '_modified.xlsx')
            wb.save(fallback_path)
            print(f"✅ فایل در این مسیر ذخیره شد: {fallback_path}")
        except Exception as e2:
            print(f"❌ خطای نهایی: {e2}")
    
except Exception as e:
    print(f"❌ خطا در پردازش فایل: {e}")
    import traceback
    traceback.print_exc()