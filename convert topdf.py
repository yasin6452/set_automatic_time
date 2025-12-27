"""
برنامه تبدیل فایل اکسل به PDF
نیاز به نصب کتابخانه‌های زیر دارد:
pip install openpyxl reportlab pandas pillow
"""

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import os
from datetime import datetime

print("📄 برنامه تبدیل فایل اکسل به PDF")
print("=" * 60)

# دریافت فایل اکسل
file_path = input("مسیر کامل فایل اکسل: ").strip().strip('"')

if not os.path.exists(file_path):
    print(f"❌ فایل پیدا نشد: {file_path}")
    exit()

print(f"✅ فایل پیدا شد: {os.path.basename(file_path)}")

try:
    # بارگذاری فایل اکسل
    wb = load_workbook(file_path, data_only=True)
    
    # نمایش لیست شیت‌ها
    print(f"\n📊 شیت‌های موجود در فایل:")
    for idx, sheet_name in enumerate(wb.sheetnames, 1):
        print(f"  {idx}. {sheet_name}")
    
    # انتخاب شیت
    sheet_choice = input(f"\n↵ شماره شیت (Enter برای شیت اول): ").strip()
    
    if sheet_choice and sheet_choice.isdigit():
        sheet_idx = int(sheet_choice) - 1
        if 0 <= sheet_idx < len(wb.sheetnames):
            ws = wb.worksheets[sheet_idx]
        else:
            print("⚠️ شماره نامعتبر، از شیت اول استفاده می‌شود")
            ws = wb.active
    else:
        ws = wb.active
    
    print(f"✅ شیت انتخاب شده: {ws.title}")
    print(f"📏 ابعاد: {ws.max_row} سطر × {ws.max_column} ستون")
    
    # تنظیمات محدوده
    print(f"\n⚙️ محدوده داده‌ها:")
    start_row = input(f"  سطر شروع (Enter برای 1): ").strip()
    start_row = int(start_row) if start_row else 1
    
    end_row = input(f"  سطر پایان (Enter برای {ws.max_row}): ").strip()
    end_row = int(end_row) if end_row else ws.max_row
    
    start_col = input(f"  ستون شروع (Enter برای 1): ").strip()
    start_col = int(start_col) if start_col else 1
    
    end_col = input(f"  ستون پایان (Enter برای {ws.max_column}): ").strip()
    end_col = int(end_col) if end_col else ws.max_column
    
    print(f"\n✅ محدوده انتخاب شده: سطر {start_row} تا {end_row}، ستون {start_col} تا {end_col}")
    
    # استخراج داده‌ها
    print(f"\n🔄 در حال استخراج داده‌ها...")
    data = []
    
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, 
                            min_col=start_col, max_col=end_col):
        row_data = []
        for cell in row:
            value = cell.value
            if value is None:
                row_data.append("")
            elif isinstance(value, datetime):
                row_data.append(value.strftime("%Y-%m-%d %H:%M"))
            else:
                row_data.append(str(value))
        data.append(row_data)
    
    print(f"✅ {len(data)} سطر استخراج شد")
    
    # تنظیمات PDF
    print(f"\n📋 تنظیمات PDF:")
    orientation = input("  جهت صفحه (1=عمودی، 2=افقی، Enter=افقی): ").strip()
    
    if orientation == "1":
        page_size = A4
        print("  ✅ جهت عمودی انتخاب شد")
    else:
        page_size = landscape(A4)
        print("  ✅ جهت افقی انتخاب شد")
    
    # نام فایل خروجی
    default_pdf_name = os.path.basename(file_path).replace('.xlsx', '.pdf').replace('.xls', '.pdf')
    output_name = input(f"\n📝 نام فایل PDF (Enter برای '{default_pdf_name}'): ").strip()
    
    if not output_name:
        output_name = default_pdf_name
    elif not output_name.endswith('.pdf'):
        output_name += '.pdf'
    
    output_path = os.path.join(os.path.dirname(file_path), output_name)
    
    # ساخت PDF
    print(f"\n🔨 در حال ساخت فایل PDF...")
    
    # ایجاد سند PDF
    doc = SimpleDocTemplate(output_path, pagesize=page_size,
                           rightMargin=30, leftMargin=30,
                           topMargin=30, bottomMargin=30)
    
    elements = []
    
    # استایل‌ها
    styles = getSampleStyleSheet()
    
    # محاسبه عرض ستون‌ها
    num_cols = len(data[0]) if data else 1
    available_width = page_size[0] - 60  # کم کردن حاشیه‌ها
    col_width = available_width / num_cols
    
    # ساخت جدول
    table = Table(data, colWidths=[col_width] * num_cols)
    
    # استایل جدول
    table_style = TableStyle([
        # سطر اول (هدر)
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # بقیه سطرها
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        
        # خطوط جدول
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BOX', (0, 0), (-1, -1), 2, colors.black),
        
        # رنگ‌بندی سطرهای زوج و فرد
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ])
    
    table.setStyle(table_style)
    elements.append(table)
    
    # ساخت PDF
    doc.build(elements)
    
    print(f"\n🎉 فایل PDF با موفقیت ساخته شد!")
    print(f"📁 مسیر: {output_path}")
    print(f"📏 حجم: {os.path.getsize(output_path):,} بایت")
    
    # خلاصه
    print(f"\n📋 خلاصه تبدیل:")
    print(f"  • فایل اکسل: {os.path.basename(file_path)}")
    print(f"  • شیت: {ws.title}")
    print(f"  • تعداد سطرها: {len(data)}")
    print(f"  • تعداد ستون‌ها: {num_cols}")
    print(f"  • جهت صفحه: {'عمودی' if page_size == A4 else 'افقی'}")
    print(f"  • فایل PDF: {os.path.basename(output_path)}")
    
except Exception as e:
    print(f"\n❌ خطا در تبدیل فایل: {e}")
    import traceback
    traceback.print_exc()

input("\n↵ برای خروج Enter بزنید...")
