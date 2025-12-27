from openpyxl import load_workbook
from datetime import time, timedelta
import random
import os
import shutil

print("🔍 برنامه پرکننده لیست کارکرد - نسخه دو شیفته")
print("=" * 60)

# دریافت فایل
file_path = input("مسیر کامل فایل اکسل: ").strip().strip('"')

if not os.path.exists(file_path):
    print(f"❌ فایل پیدا نشد: {file_path}")
    exit()

print(f"✅ فایل پیدا شد: {os.path.basename(file_path)}")

# ایجاد پشتیبان
backup_path = file_path.replace('.xlsx', '_backup.xlsx')
shutil.copy2(file_path, backup_path)
print(f"📋 پشتیبان گرفته شد: {os.path.basename(backup_path)}")

try:
    # بارگذاری فایل
    wb = load_workbook(file_path)
    
    # انتخاب شیت صحیح
    if 'مریم ایرانمنش' in wb.sheetnames:
        ws = wb['مریم ایرانمنش']
        print(f"✅ شیت 'مریم ایرانمنش' انتخاب شد")
    else:
        ws = wb.active
        print(f"⚠️ شیت فعال: {ws.title}")
    
    print(f"📊 ابعاد شیت: {ws.max_row} سطر × {ws.max_column} ستون")
    
    # بررسی ساختار فایل
    print("\n🔍 بررسی ساختار فایل...")
    
    # پیدا کردن ستون‌های کلیدی
    entry_col_shift2 = None  # ستون ورود شیفت دوم
    exit_col_shift2 = None   # ستون خروج شیفت دوم
    work_col = None          # ستون کارکرد
    day_col = None           # ستون روز
    
    # بررسی سطر 7 و 8 برای هدرها
    for col in range(1, ws.max_column + 1):
        val7 = str(ws.cell(row=7, column=col).value or "")
        val8 = str(ws.cell(row=8, column=col).value or "")
        
        if "ترددها" in val8 and "ورود" in val7:
            entry_col_shift2 = col
            print(f"✅ ستون ورود شیفت 2 پیدا شد: {chr(64 + col)} (ستون {col})")
        elif "ترددها" in val8 and "خروج" in val7:
            exit_col_shift2 = col
            print(f"✅ ستون خروج شیفت 2 پیدا شد: {chr(64 + col)} (ستون {col})")
        elif "طول" in val7 or "كاركرد" in val8:
            work_col = col
            print(f"✅ ستون کارکرد پیدا شد: {chr(64 + col)} (ستون {col})")
        elif "روز" in val7 or "روز" in val8:
            day_col = col
            print(f"✅ ستون روز پیدا شد: {chr(64 + col)} (ستون {col})")
    
    # اگر ستون روز پیدا نشد، از ستون 35 استفاده کن
    if not day_col:
        day_col = 35
        print(f"⚠️ ستون روز پیدا نشد، از ستون {day_col} استفاده می‌کنم")
    
    # اگر ستون‌های شیفت 2 پیدا نشدند
    if not entry_col_shift2:
        entry_col_shift2 = 27  # AA
        print(f"⚠️ ستون ورود شیفت 2 پیدا نشد، از ستون {entry_col_shift2} استفاده می‌کنم")
    
    if not exit_col_shift2:
        exit_col_shift2 = 28  # AB
        print(f"⚠️ ستون خروج شیفت 2 پیدا نشد، از ستون {exit_col_shift2} استفاده می‌کنم")
    
    if not work_col:
        work_col = 26  # Z
        print(f"⚠️ ستون کارکرد پیدا نشد، از ستون {work_col} استفاده می‌کنم")
    
    # تعیین ستون‌های شیفت 1 (صبح)
    # ستون‌های AF=32 و AE=31
    entry_col_shift1 = 31  # AE
    exit_col_shift1 = 32   # AF
    
    print(f"\n🎯 ستون‌های نهایی:")
    print(f"  • ستون روزها: AI (شماره {day_col})")
    print(f"  • ستون ورود شیفت 1 (صبح): AE (شماره {entry_col_shift1})")
    print(f"  • ستون خروج شیفت 1 (صبح): AF (شماره {exit_col_shift1})")
    
    # تبدیل شماره ستون به حروف برای نمایش بهتر
    def col_num_to_letter(col_num):
        letter = ""
        while col_num > 0:
            col_num -= 1
            letter = chr(col_num % 26 + 65) + letter
            col_num //= 26
        return letter
    
    print(f"  • ستون ورود شیفت 2 (عصر): {col_num_to_letter(entry_col_shift2)} (شماره {entry_col_shift2})")
    print(f"  • ستون خروج شیفت 2 (عصر): {col_num_to_letter(exit_col_shift2)} (شماره {exit_col_shift2})")
    print(f"  • ستون کارکرد: {col_num_to_letter(work_col)} (شماره {work_col})")
    
    # تنظیمات شیفت‌ها
    print(f"\n⏰ شیفت اول (صبح):")
    print(f"  • ورود: 5:00، 5:10 (رندوم)")
    print(f"  • خروج: 11:00، 11:10، 11:20 (رندوم)")
    
    print(f"\n⏰ شیفت دوم (عصر):")
    print(f"  • ورود: 14:50، 15:00، 15:10 (رندوم)")
    print(f"  • خروج: 18:00، 18:10 (رندوم)")
    
    # تایید از کاربر
    confirm = input("\n↵ برای ادامه Enter بزنید (یا 'n' برای لغو): ").strip().lower()
    if confirm == 'n':
        print("❌ عملیات لغو شد.")
        exit()
    
    print("\n🔄 در حال پر کردن داده‌ها...")
    
    filled_days = 0
    fridays = 0
    
    # گزینه‌های زمانی
    shift1_entry_options = [
        time(5, 0),
        time(5, 10)
    ]
    
    shift1_exit_options = [
        time(11, 0),
        time(11, 10),
        time(11, 20)
    ]
    
    shift2_entry_options = [
        time(14, 50),
        time(15, 0),
        time(15, 10)
    ]
    
    shift2_exit_options = [
        time(18, 0),
        time(18, 10)
    ]
    
    for row in range(9, 40):  # سطر 9 تا 39
        day_value = ws.cell(row=row, column=day_col).value
        
        if not day_value:
            continue
            
        day_str = str(day_value).strip()
        
        if day_str == "جمعه":
            # جمعه: پاک کردن
            ws.cell(row=row, column=entry_col_shift1).value = None
            ws.cell(row=row, column=exit_col_shift1).value = None
            ws.cell(row=row, column=entry_col_shift2).value = None
            ws.cell(row=row, column=exit_col_shift2).value = None
            ws.cell(row=row, column=work_col).value = None
            fridays += 1
            print(f"🚫 سطر {row:2d} (جمعه) - پاک شد")
        else:
            # روز کاری: پر کردن دو شیفت
            
            # شیفت اول (صبح)
            shift1_entry = random.choice(shift1_entry_options)
            shift1_exit = random.choice(shift1_exit_options)
            
            # محاسبه مدت شیفت اول (به دقیقه)
            shift1_minutes = (shift1_exit.hour * 60 + shift1_exit.minute) - (shift1_entry.hour * 60 + shift1_entry.minute)
            
            # شیفت دوم (عصر)
            shift2_entry = random.choice(shift2_entry_options)
            shift2_exit = random.choice(shift2_exit_options)
            
            # محاسبه مدت شیفت دوم (به دقیقه)
            shift2_minutes = (shift2_exit.hour * 60 + shift2_exit.minute) - (shift2_entry.hour * 60 + shift2_entry.minute)
            
            # مجموع کارکرد (به ساعت)
            total_work_hours = (shift1_minutes + shift2_minutes) / 60
            
            # پر کردن سلول‌ها - شیفت اول
            ws.cell(row=row, column=entry_col_shift1).value = shift1_entry
            ws.cell(row=row, column=exit_col_shift1).value = shift1_exit
            ws.cell(row=row, column=entry_col_shift1).number_format = "HH:MM"
            ws.cell(row=row, column=exit_col_shift1).number_format = "HH:MM"
            
            # پر کردن سلول‌ها - شیفت دوم
            ws.cell(row=row, column=entry_col_shift2).value = shift2_entry
            ws.cell(row=row, column=exit_col_shift2).value = shift2_exit
            ws.cell(row=row, column=entry_col_shift2).number_format = "HH:MM"
            ws.cell(row=row, column=exit_col_shift2).number_format = "HH:MM"
            
            # پر کردن کارکرد کل
            ws.cell(row=row, column=work_col).value = round(total_work_hours, 1)
            ws.cell(row=row, column=work_col).number_format = "0.0"
            
            filled_days += 1
            print(f"✅ سطر {row:2d} ({day_str}):")
            print(f"   شیفت 1: {shift1_entry.strftime('%H:%M')} - {shift1_exit.strftime('%H:%M')} ({shift1_minutes // 60}:{shift1_minutes % 60:02d})")
            print(f"   شیفت 2: {shift2_entry.strftime('%H:%M')} - {shift2_exit.strftime('%H:%M')} ({shift2_minutes // 60}:{shift2_minutes % 60:02d})")
            print(f"   کل: {total_work_hours:.1f} ساعت")
    
    print(f"\n📊 نتیجه نهایی:")
    print(f"  • {filled_days} روز کاری پر شد")
    print(f"  • {fridays} جمعه پاک شد")
    print(f"  • شیفت اول: 5:00-5:10 تا 11:00-11:20")
    print(f"  • شیفت دوم: 14:50-15:10 تا 18:00-18:10")
    
    # ذخیره فایل
    default_name = os.path.basename(file_path).replace('.xlsx', '_دو_شیفته.xlsx')
    output_name = input(f"\n📝 نام فایل خروجی (Enter برای '{default_name}'): ").strip()
    
    if not output_name:
        output_name = default_name
    elif not output_name.endswith('.xlsx'):
        output_name += '.xlsx'
    
    output_path = os.path.join(os.path.dirname(file_path), output_name)
    
    try:
        wb.save(output_path)
        print(f"\n🎉 فایل با موفقیت ذخیره شد!")
        print(f"📁 مسیر: {output_path}")
        print(f"📏 حجم: {os.path.getsize(output_path):,} بایت")
        
        # نمایش خلاصه
        print(f"\n📋 خلاصه تغییرات:")
        print(f"  1. شیفت اول: ورود 5:00-5:10، خروج 11:00-11:20")
        print(f"  2. شیفت دوم: ورود 14:50-15:10، خروج 18:00-18:10")
        print(f"  3. استفاده از 4 ستون (2 ستون برای هر شیفت)")
        print(f"  4. کارکرد کل = مجموع دو شیفت")
        print(f"  5. روزهای جمعه خالی شدند")
        print(f"  6. فرمت زمان: HH:MM")
        print(f"  7. فرمت کارکرد: 0.0 ساعت")
        
    except Exception as e:
        print(f"❌ خطا در ذخیره فایل: {e}")
        print("⚠️ در حال ذخیره با نام پیش‌فرض...")
        try:
            fallback_path = file_path.replace('.xlsx', '_modified.xlsx')
            wb.save(fallback_path)
            print(f"✅ فایل در این مسیر ذخیره شد: {fallback_path}")
        except Exception as e2:
            print(f"❌ خطای نهایی: {e2}")
    
except Exception as e:
    print(f"❌ خطا در پردازش فایل: {e}")
    import traceback
    traceback.print_exc()